from io import BytesIO
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Border, Side
import streamlit as st

def create_combined_excel(df1, df2, discrepancy_indices1, discrepancy_indices2, selected_keys, discrepancies):
    """
    Create a combined Excel file with both dataframes in separate sheets with highlighted discrepancies
    """
    output = BytesIO()
    
    # Debug print
    print("Value mismatches:", discrepancies['value_mismatches'])
    print("Missing in df2:", discrepancies['missing_rows']['missing_in_df2'])
    print("Missing in df1:", discrepancies['missing_rows']['missing_in_df1'])
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Reorder columns to put selected_keys first
        cols1 = df1.columns.tolist()
        cols2 = df2.columns.tolist()
        
        # Remove selected_keys from current position and add it to the front
        if selected_keys in cols1:
            cols1.remove(selected_keys)
            cols1.insert(0, selected_keys)
            df1 = df1[cols1]
        
        if selected_keys in cols2:
            cols2.remove(selected_keys)
            cols2.insert(0, selected_keys)
            df2 = df2[cols2]
            
        # Write dataframes to Excel
        df1.to_excel(writer, sheet_name='File 1', index=False)
        df2.to_excel(writer, sheet_name='File 2', index=False)
        
        # Get workbook and sheets
        workbook = writer.book
        worksheet1 = writer.sheets['File 1']
        worksheet2 = writer.sheets['File 2']
        
        # Define highlight styles with more vibrant colors
        value_mismatch_style = PatternFill(
            start_color='FF9999',  # Light red
            end_color='FF9999',
            fill_type='solid'
        )
        missing_row_style = PatternFill(
            start_color='FFB366',  # Light orange
            end_color='FFB366',
            fill_type='solid'
        )

        # Helper function to safely convert to string and strip
        def safe_strip(value):
            if isinstance(value, tuple):
                return str(value[0]).strip()
            return str(value).strip()

        # Create lookup dictionaries for mismatches
        value_mismatches = {}
        for mismatch in discrepancies['value_mismatches']:
            key = safe_strip(mismatch['key'])
            if key not in value_mismatches:
                value_mismatches[key] = []
            value_mismatches[key].append(safe_strip(mismatch['column']))

        # Create sets for missing rows
        missing_in_df2 = {safe_strip(k) for k in discrepancies['missing_rows']['missing_in_df2']}
        missing_in_df1 = {safe_strip(k) for k in discrepancies['missing_rows']['missing_in_df1']}

        # Helper function to get column index
        def get_column_index(worksheet, column_name):
            for idx, cell in enumerate(worksheet[1]):
                if safe_strip(cell.value) == safe_strip(column_name):
                    return idx
            return None

        # Highlight discrepancies in first sheet
        for row_idx in range(len(df1)):
            key = safe_strip(df1.iloc[row_idx][selected_keys])
            
            # Handle value mismatches - highlight only mismatched cells
            if key in value_mismatches:
                for mismatched_col in value_mismatches[key]:
                    col_idx = get_column_index(worksheet1, mismatched_col)
                    if col_idx is not None:
                        cell = worksheet1.cell(row=row_idx + 2, column=col_idx + 1)
                        cell.fill = value_mismatch_style
            
            # Handle missing rows - highlight entire row
            if key in missing_in_df2:
                for col_idx in range(len(df1.columns)):
                    cell = worksheet1.cell(row=row_idx + 2, column=col_idx + 1)
                    cell.fill = missing_row_style

        # Highlight discrepancies in second sheet
        for row_idx in range(len(df2)):
            key = safe_strip(df2.iloc[row_idx][selected_keys])
            
            # Handle value mismatches - highlight only mismatched cells
            if key in value_mismatches:
                for mismatched_col in value_mismatches[key]:
                    col_idx = get_column_index(worksheet2, mismatched_col)
                    if col_idx is not None:
                        cell = worksheet2.cell(row=row_idx + 2, column=col_idx + 1)
                        cell.fill = value_mismatch_style
            
            # Handle missing rows - highlight entire row
            if key in missing_in_df1:
                for col_idx in range(len(df2.columns)):
                    cell = worksheet2.cell(row=row_idx + 2, column=col_idx + 1)
                    cell.fill = missing_row_style

        # Auto-adjust column widths
        for worksheet in [worksheet1, worksheet2]:
            for column in worksheet.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[column[0].column_letter].width = adjusted_width

        # Update legend text to include key column information
        legend_text = [
            ["Legend:"],
            ["Key Column:", selected_keys],  # Add key column information
            ["Red", "Value mismatch between files"],
            ["Orange", "Row missing in other file"]
        ]
        
        for worksheet in [worksheet1, worksheet2]:
            start_row = len(df1 if worksheet == worksheet1 else df2) + 4
            
            # Add legend with borders
            for i, row in enumerate(legend_text):
                for j, text in enumerate(row):
                    cell = worksheet.cell(row=start_row + i, column=j + 1)
                    cell.value = text
                    cell.border = openpyxl.styles.Border(
                        left=openpyxl.styles.Side(style='thin'),
                        right=openpyxl.styles.Side(style='thin'),
                        top=openpyxl.styles.Side(style='thin'),
                        bottom=openpyxl.styles.Side(style='thin')
                    )
                    if i > 1 and j == 0:  # Color cells in first column of rows 3 and 4 (shifted down by 1 due to new key info)
                        cell.fill = value_mismatch_style if i == 2 else missing_row_style
    
    output.seek(0)
    return output

def get_displayed_dataframe(df):
    """
    Returns the currently displayed dataframe based on applied filters and sorting
    
    Args:
        df (pd.DataFrame): Original dataframe
        
    Returns:
        pd.DataFrame: Filtered and sorted dataframe
    """
    # If you're using session state to store filtered data, return that
    if 'filtered_df' in st.session_state:
        return st.session_state.filtered_df
    
    # If not using session state, return the original dataframe
    return df 